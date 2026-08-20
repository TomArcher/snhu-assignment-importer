from copy import copy
import re

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.table import Table
from playwright.sync_api import sync_playwright
from playwright.sync_api import Error as PlaywrightError

# Constants for SNHU
SNHU_URL = "https://learn.snhu.edu/"
TERM_FORMULA_SECOND_ROW = "=VLOOKUP($B2,CourseData[#Data],2,FALSE)"
         
# Constants for instructions
instructions = """
*** INSTRUCTIONS ***
From the browser that this app opens, log in to SNHU.
Navigate to the desired Grades page.
Return here.
"""

get_course_name = "Specify the course name (ex: CS-350) or Q to quit: "

# Constants for the Excel workbook and Data tab names
excel_workbook_name = "snhu_template.xlsx"
data_tab_name = "Data"


def get_assignments(page) -> list[dict]:
    """Extract assignments from the current SNHU Grades page.

    Args:
        page: Playwright page representing the current browser page.

    Returns:
        list[dict]: Represents a list of assignments where each
        assignment is a dictionary of values (containing values
        for the module number, assignment name, and max points).
    """

    # Get the text from the current browser page.
    text = page.locator("body").inner_text()

    # Create a list containing each non-empty line from the browser
    # page and remove any leading or trailing whitespace.
    lines = [
        line.strip()
        for line in text.splitlines()
        if line.strip()
    ]

    # Initialize the assignments list
    assignments = []

    # For every line in the browser page...
    for i, line in enumerate(lines):

        # Each assignment grade row has a "Points" column with
        # the format "- / xx" (where xx is the maximum points).
        # Use regex to determine if the current line contains the
        # value extracted from this column.
        match = re.fullmatch(r"- / (\d+)", line)

        # If the current line contains the maximum points
        # (a regex match was found) and a preceding line exists for
        # the assignment name...
        if match and i > 0:

            # Get the assignment name
            assignment_name = lines[i - 1]

            # Get the max points
            max_points = int(match.group(1))

            # Each assignment name begins with an "x-y" prefix, where x
            # is the module number and y is the assignment number. Use
            # regex to match this prefix and capture the module number.
            module_match = re.match(r"(\d+)-\d+", assignment_name)

            # If a regex match was found...
            if module_match:

                # Get the module number from the module match
                module = int(module_match.group(1))

            else:  # If a match was NOT found...

                # Set module to None to indicate that it was not
                # found. This code is defensive and should never
                # run.
                module = None

            # Create a dictionary containing the current assignment's
            # module number, assignment name, and maximum points, and
            # append it to the assignments list.
            assignments.append({
                "Module": module,
                "Assignment": assignment_name,
                "Max": max_points
            })

    # Return the list of assignments extracted from the current
    # SNHU Grades page.
    return assignments


def find_course(workbook, course_name):
    """Find a course in the workbook's course data table.

    Args:
        workbook: The openpyxl workbook containing the course data.
        course_name: The name of the course to find (ex: "CS-350").

    Returns:
        dict: Represents the course information if the specified course
        was found. Returns None if the course was not found.
    """

    # Get the worksheet containing the course data.
    data_sheet = workbook[data_tab_name]

    # Create a list containing the column headings from the first
    # row of the course data worksheet.
    headers = [
        cell.value
        for cell in data_sheet[1]
    ]

    # Initialize the course to None to indicate that the specified
    # course has not yet been found.
    course = None

    # For every course row in the course data worksheet, starting
    # with row 2 to exclude the column headings...
    for row in data_sheet.iter_rows(
        min_row=2,
        values_only=True
    ):

        # Create a dictionary containing the current course information
        # by pairing each column heading with its corresponding value.
        current_course = dict(zip(headers, row))

        # If the current course name matches the specified course name...
        if current_course["Name"] == course_name:

            # Set the course to the current course information.
            course = current_course

            # Stop searching because the specified course was found.
            break

    # Return the course information if the specified course was found,
    # or None if the course was not found.
    return course


def display_course(course):
    """Display information for the specified course.

    Args:
        course: Dictionary containing the course information.
    """

    # Display the course name, term, and description.
    print("\nCourse found:")
    print(f"Name:        {course['Name']}")
    print(f"Term:        {course['Term']}")
    print(f"Description: {course['Description']}")


def get_grades_context(workbook):
    """Get the Grades worksheet, table, and table boundaries.

    Args:
        workbook: The openpyxl workbook containing the Grades worksheet.

    Returns:
        tuple: Contains the Grades worksheet, Grades table, and a tuple
        representing the table's minimum and maximum rows and columns.
    """

    # Get the Grades worksheet from the workbook.
    grades_sheet = workbook["Grades"]

    # Get the Excel table containing the Grades data.
    grades_table: Table = grades_sheet.tables["Table3"]

    # Get the minimum and maximum rows and columns occupied by the
    # Grades table.
    bounds = range_boundaries(grades_table.ref)

    # Return the Grades worksheet, Grades table, and table boundaries.
    return grades_sheet, grades_table, bounds


def find_existing_assignments(
    grades_sheet,
    course_name,
    min_row,
    totals_row
):
    """Find assignments already stored for the specified course.

    Args:
        grades_sheet: Worksheet containing the Grades table.
        course_name: The name of the course to search for.
        min_row: The first row occupied by the Grades table.
        totals_row: The row containing the Grades table totals.

    Returns:
        list: Represents the assignment names already stored for the
        specified course.
    """

    # Initialize the list of existing assignments.
    existing_assignments = []

    # For every data row in the Grades table, excluding the header
    # and totals rows...
    for row_number in range(min_row + 1, totals_row):

        # If the current row belongs to the specified course...
        if grades_sheet.cell(row_number, 2).value == course_name:

            # Add the current assignment name to the list of existing
            # assignments for the specified course.
            existing_assignments.append(
                grades_sheet.cell(row_number, 5).value
            )

    # Return the assignments already stored for the specified course.
    return existing_assignments


def preview_assignments(course, assignments):
    """Display the assignments that will be imported.

    Args:
        course: Dictionary containing the course information.
        assignments: List of assignment dictionaries to be imported.
    """

    # Display the number of assignments found on the SNHU Grades page.
    print(f"\nAssignments found: {len(assignments)}")

    # Calculate and display the total number of possible points for
    # all assignments that will be imported.
    print(
        "Total possible points: "
        f"{sum(assignment['Max'] for assignment in assignments)}"
    )

    if len(assignments) > 0:
        
        # Display a heading for the assignment rows that will be added.
        print("\nRows that will be added:")

        # For every assignment that will be imported...
        for assignment in assignments:

            # Display the course name, module number, assignment name,
            # and maximum points for the current assignment.
            print(
                course["Name"],
                assignment["Module"],
                assignment["Assignment"],
                assignment["Max"]
            )

    return len(assignments)

def confirm_import():
    """Ask the user whether the assignments should be imported.

    Returns:
        bool: True if the user confirms the import; otherwise False.
    """

    # Ask the user to confirm whether the assignments should be
    # added to the Excel workbook.
    answer = input(
        f"\nAdd these assignments to {excel_workbook_name}? (y/n): "
    ).strip().lower()

    # Return True only if the user entered "y".
    return answer == "y"

def update_grades_formulas(grades_sheet, totals_row):
    """Update the formulas for each assignment row in the Grades table."""

    # For every assignment row in the Grades table...
    for row in range(2, totals_row):

        # Set the Term formula to reference the course name in the
        # current row.
        grades_sheet.cell(row=row, column=1).value = (
            f'=VLOOKUP($B{row},CourseData[#Data],2,FALSE)'
        )

        # Set the Description formula to reference the course name
        # in the current row.
        grades_sheet.cell(row=row, column=3).value = (
            f'=VLOOKUP($B{row},CourseData[#Data],6,FALSE)'
        )

        # Set the Grade formula to calculate the grade using the
        # Points and Max values in the current row.
        grades_sheet.cell(row=row, column=8).value = (
            f'=IF(ISNUMBER(G{row}),G{row}/F{row},"")'
        )

def delete_placeholder_row(grades_sheet, grades_table):
    """If it exists, delete the placeholder row."""

    # If the first row and column do not contain the formula that
    # exists in every imported row, the row must be the placeholder
    # row that Excel requires for an otherwise empty table...
    if grades_sheet.cell(row=2, column=1).value != TERM_FORMULA_SECOND_ROW:

        # Get the current boundaries of the Grades table.
        min_col, min_row, max_col, max_row = range_boundaries(
            grades_table.ref
        )

        # Delete Excel's required empty placeholder row.
        grades_sheet.delete_rows(2)

        # The deletion moved the Total row up one row, so shrink the
        # Excel table range to match the worksheet.
        new_totals_row = max_row - 1

        grades_table.ref = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{new_totals_row}"
        )

        # If the table contains an AutoFilter...
        if grades_table.autoFilter is not None:

            # Shrink the AutoFilter range to exclude the Total row.
            grades_table.autoFilter.ref = (
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(max_col)}{new_totals_row - 1}"
            )

        # Update the formulas because openpyxl does not adjust formula
        # references when worksheet rows are deleted.
        update_grades_formulas(
            grades_sheet,
            new_totals_row
        )

def insert_assignments(
    grades_sheet,
    grades_table,
    course,
    assignments,
    min_col,
    min_row,
    max_col,
    max_row
):
    """Insert assignments into the Grades table.

    Args:
        grades_sheet: Worksheet containing the Grades table.
        grades_table: Excel table containing the Grades data.
        course: Dictionary containing the course information.
        assignments: List of assignment dictionaries to insert.
        min_col: The first column occupied by the Grades table.
        min_row: The first row occupied by the Grades table.
        max_col: The last column occupied by the Grades table.
        max_row: The last row occupied by the Grades table.
    """

    # The last row in the Grades table contains the totals row.
    totals_row = max_row

    # Use the last normal data row immediately above the totals row as
    # the template for formatting and formulas in the new rows.
    template_row = totals_row - 1

    # Get the number of rows that need to be inserted into the Grades
    # table from the number of assignments being imported.
    number_of_assignments = len(assignments)

    # Insert enough blank worksheet rows immediately before the totals
    # row to hold all assignments being imported.
    grades_sheet.insert_rows(
        totals_row,
        amount=number_of_assignments
    )

    # For every assignment being imported...
    for offset, assignment in enumerate(assignments):

        # Calculate the worksheet row where the current assignment
        # should be inserted.
        target_row = totals_row + offset

        # Copy the formatting from the template row to the row where
        # the current assignment will be inserted.
        copy_row_formatting(
            grades_sheet,
            template_row,
            target_row,
            min_col,
            max_col
        )

        # Copy and translate the formulas from the template row to the
        # row where the current assignment will be inserted.
        copy_row_formulas(
            grades_sheet,
            template_row,
            target_row
        )

        # Write the current assignment values to the target row.
        write_assignment_row(
            grades_sheet,
            target_row,
            course,
            assignment
        )

    # Expand the Grades table to include all newly inserted rows and
    # the totals row at its new location.
    update_grades_table(
        grades_table,
        min_col,
        min_row,
        max_col,
        max_row + number_of_assignments
    )

    delete_placeholder_row(grades_sheet, grades_table)

def copy_row_formatting(
    grades_sheet,
    source_row,
    target_row,
    min_col,
    max_col
):
    """Copy cell formatting from one Grades row to another.

    Args:
        grades_sheet: Worksheet containing the Grades table.
        source_row: Row containing the formatting to copy.
        target_row: Row that will receive the copied formatting.
        min_col: The first column whose formatting should be copied.
        max_col: The last column whose formatting should be copied.
    """

    # For every column in the Grades table...
    for column in range(min_col, max_col + 1):

        # Get the source cell containing the formatting to copy.
        source_cell = grades_sheet.cell(source_row, column)

        # Get the corresponding target cell that will receive the
        # copied formatting.
        target_cell = grades_sheet.cell(target_row, column)

        # If the source cell has a style...
        if source_cell.has_style:

            # Copy the complete cell style to the target cell.
            target_cell._style = copy(source_cell._style)

        # If the source cell has a number format...
        if source_cell.number_format:

            # Copy the number format to the target cell.
            target_cell.number_format = source_cell.number_format

        # Copy the remaining formatting properties from the source
        # cell to the target cell.
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)


def copy_row_formulas(grades_sheet, source_row, target_row):
    """Copy and translate the calculated formulas to a new row.

    Args:
        grades_sheet: Worksheet containing the Grades table.
        source_row: Row containing the formulas to copy.
        target_row: Row that will receive the translated formulas.
    """

    # For every calculated column in the Grades table...
    for column in (1, 3, 8):

        # Get the source cell containing the formula to copy.
        source_cell = grades_sheet.cell(source_row, column)

        # Get the corresponding target cell that will receive the
        # translated formula.
        target_cell = grades_sheet.cell(target_row, column)

        # If the source cell contains an Excel formula...
        if (
            isinstance(source_cell.value, str)
            and source_cell.value.startswith("=")
        ):

            # Translate the source formula so its relative references
            # point to the target row, and store it in the target cell.
            target_cell.value = Translator(
                source_cell.value,
                origin=source_cell.coordinate
            ).translate_formula(target_cell.coordinate)


def write_assignment_row(
    grades_sheet,
    target_row,
    course,
    assignment
):
    """Write assignment values into the specified Grades row.

    Args:
        grades_sheet: Worksheet containing the Grades table.
        target_row: Row where the assignment values should be written.
        course: Dictionary containing the course information.
        assignment: Dictionary containing the assignment information.
    """

    # Write the calculated Term formula.
    grades_sheet.cell(target_row, 1).value = (
        f'=VLOOKUP($B{target_row},CourseData[#Data],2,FALSE)'
    )

    # Write the course name to the Name column.
    grades_sheet.cell(target_row, 2).value = course["Name"]

    # Write the calculated Description formula.
    grades_sheet.cell(target_row, 3).value = (
        f'=VLOOKUP($B{target_row},CourseData[#Data],6,FALSE)'
    )

    # Write the module number to the Module column.
    grades_sheet.cell(target_row, 4).value = assignment["Module"]

    # Write the assignment name to the Assignment column.
    grades_sheet.cell(target_row, 5).value = assignment["Assignment"]

    # Write the maximum possible points to the Max column.
    grades_sheet.cell(target_row, 6).value = assignment["Max"]

    # Leave the Points column blank because no grade has been received yet.
    grades_sheet.cell(target_row, 7).value = None

    # Write the calculated Grade formula.
    grades_sheet.cell(target_row, 8).value = (
        f'=IF(ISNUMBER(G{target_row}),G{target_row}/F{target_row},"")'
    )

def update_grades_table(
    grades_table,
    min_col,
    min_row,
    max_col,
    new_totals_row
):
    """Expand the Grades table and AutoFilter to include the new rows.

    Args:
        grades_table: Excel table containing the Grades data.
        min_col: The first column occupied by the Grades table.
        min_row: The first row occupied by the Grades table.
        max_col: The last column occupied by the Grades table.
        new_totals_row: The new worksheet row containing the totals row.
    """

    # Expand the Grades table reference through the relocated totals row.
    grades_table.ref = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{new_totals_row}"
    )

    # If the Grades table has an AutoFilter...
    if grades_table.autoFilter is not None:

        # Expand the AutoFilter through the last data row. The totals row
        # is intentionally excluded from the AutoFilter range.
        grades_table.autoFilter.ref = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{new_totals_row - 1}"
        )


def import_assignments(workbook, course, assignments):
    """Import assignments into the workbook's Grades table.

    Args:
        workbook: The openpyxl workbook containing the Grades table.
        course: Dictionary containing the course information.
        assignments: List of assignment dictionaries to import.
    """

    # Display information for the course being imported.
    display_course(course)

    # Get the Grades worksheet, Excel table, and table boundaries.
    grades_sheet, grades_table, bounds = get_grades_context(workbook)

    # Unpack the Grades table boundaries into their individual values.
    min_col, min_row, max_col, max_row = bounds

    # The last row in the Grades table contains the totals row.
    totals_row = max_row

    # Find any assignments that have already been imported for the
    # specified course.
    existing_assignments = find_existing_assignments(
        grades_sheet,
        course["Name"],
        min_row,
        totals_row
    )

    # If assignments already exist for the specified course...
    if existing_assignments:

        # Display how many assignments already exist for the course.
        print(
            f"\n{course['Name']} already has "
            f"{len(existing_assignments)} assignments in Grades.\n"
        )

        # Cancel the import to prevent duplicate assignment rows.
        print("Import cancelled to prevent duplicates.\n")
        return

    # Display the assignments that will be imported.
    assignments_count = preview_assignments(course, assignments)

    if assignments_count > 0:
        # If the user does NOT confirm the import...
        if not confirm_import():

            # Cancel the import without changing the workbook.
            print("\nImport cancelled.\n")
            return

        # Insert the assignments into the Grades table.
        insert_assignments(
            grades_sheet,
            grades_table,
            course,
            assignments,
            min_col,
            min_row,
            max_col,
            max_row
        )

        # Save the updated workbook to the original Excel file.
        workbook.save(excel_workbook_name)

        print(
            f"\nAdded {len(assignments)} assignments "
            f"and saved them to {excel_workbook_name}.\n"
        )
    else:
        print(
            f"\nNo assignments to add to spreadsheet.\n"
        )


def main():
    """Run the SNHU grade assignment importer."""

    # Start Playwright and automatically clean up its resources when
    # the application exits the with block.
    with sync_playwright() as p:

        # Load the Excel workbook that contains the course and grade data.
        workbook = load_workbook(excel_workbook_name, data_only=False)

        # Launch a persistent Chromium browser context so the SNHU login
        # session can be reused between application runs.
        context = p.chromium.launch_persistent_context(
            user_data_dir="browser-profile",
            headless=False
        )

        # Get the browser page opened by the persistent Chromium context.
        page = context.pages[0]

        # Navigate the browser to the SNHU learning site.
        page.goto(SNHU_URL)

        # Display the instructions for navigating to a course Grades page.
        print(instructions)

        # Continue importing courses until the user chooses to quit...
        while True:

            # Ask the user for the course name associated with the Grades
            # page currently displayed in the browser.
            course_name = input(get_course_name).strip().upper()

            # If the user entered "Q"...
            if course_name == "Q":

                # Exit the course import loop.
                break

            # Extract the assignments from the current SNHU Grades page.
            assignments = get_assignments(page)

            # Find the specified course in the workbook's course data.
            course = find_course(workbook, course_name)

            # If the specified course was NOT found...
            if course is None:

                # Display an error and allow the user to try another course.
                print(
                    f"\nCourse {course_name} was not found in CourseData."
                )

            else:  # If the specified course was found...

                # Import the assignments into the workbook's Grades table.
                import_assignments(workbook, course, assignments)

        # Close the persistent browser context before exiting the app.
        try:
            context.close()
        except PlaywrightError:
            pass

# If this file is being run directly rather than imported as a module...
if __name__ == "__main__":

    # Run the application.
    main()
